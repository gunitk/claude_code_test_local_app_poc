"""
Excel export utilities for test cases and execution results
This module provides functionality to export test data to Excel format
"""

import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelExporter:
    """Utility class for exporting test data to Excel format"""
    
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        self.wrap_alignment = Alignment(wrap_text=True, vertical="top")
    
    def export_test_cases(self, test_cases_data, session_id, filename=None):
        """
        Export test cases to Excel format
        
        Args:
            test_cases_data (list): List of test case dictionaries
            session_id (str): Session identifier
            filename (str, optional): Output filename
        
        Returns:
            str: Path to the created Excel file
        """
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"test_cases_{session_id}_{timestamp}.xlsx"
        
        filepath = f"downloads/{session_id}/{filename}"
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        
        # Define headers
        headers = [
            "ID", "Name", "Description", "Priority", "Category", 
            "Steps", "Expected Result", "Test Data", "Estimated Time"
        ]
        
        # Add headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Add test case data
        for row_num, test_case in enumerate(test_cases_data, 2):
            # Convert steps list to numbered string
            steps_text = ""
            if isinstance(test_case.get('steps'), list):
                steps_text = "\n".join([f"{i+1}. {step}" for i, step in enumerate(test_case['steps'])])
            else:
                steps_text = str(test_case.get('steps', ''))
            
            # Convert test_data dict to formatted string
            test_data_text = ""
            if isinstance(test_case.get('test_data'), dict) and test_case['test_data']:
                test_data_text = "\n".join([f"{k}: {v}" for k, v in test_case['test_data'].items()])
            else:
                test_data_text = str(test_case.get('test_data', ''))
            
            # Populate row data
            row_data = [
                test_case.get('id', row_num - 1),
                test_case.get('name', ''),
                test_case.get('description', ''),
                test_case.get('priority', ''),
                test_case.get('category', ''),
                steps_text,
                test_case.get('expected_result', ''),
                test_data_text,
                test_case.get('estimated_time', '')
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = self.border
                cell.alignment = self.wrap_alignment
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
        # Add summary information
        self._add_summary_sheet(wb, test_cases_data, session_id)
        
        # Save workbook
        wb.save(filepath)
        return filepath
    
    def export_execution_results(self, execution_data, session_id, filename=None):
        """
        Export test execution results to Excel format
        
        Args:
            execution_data (dict): Execution results data
            session_id (str): Session identifier
            filename (str, optional): Output filename
        
        Returns:
            str: Path to the created Excel file
        """
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"execution_results_{session_id}_{timestamp}.xlsx"
        
        filepath = f"downloads/{session_id}/{filename}"
        
        # Create workbook
        wb = Workbook()
        
        # Create execution results sheet
        ws_results = wb.active
        ws_results.title = "Execution Results"
        
        # Headers for execution results
        headers = [
            "Test ID", "Test Name", "Status", "Execution Time", 
            "Error Message", "Screenshot", "Execution Details"
        ]
        
        # Add headers
        for col_num, header in enumerate(headers, 1):
            cell = ws_results.cell(row=1, column=col_num, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Get execution results
        results = execution_data.get('results', [])
        
        # Add execution data
        for row_num, result in enumerate(results, 2):
            # Format execution details
            details = ""
            if result.get('details'):
                details = "\n".join([f"{k}: {v}" for k, v in result['details'].items()])
            
            row_data = [
                result.get('test_id', ''),
                result.get('test_name', ''),
                result.get('status', ''),
                result.get('execution_time', ''),
                result.get('error_message', ''),
                result.get('screenshot', ''),
                details
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws_results.cell(row=row_num, column=col_num, value=value)
                cell.border = self.border
                cell.alignment = self.wrap_alignment
                
                # Color code status cells
                if col_num == 3:  # Status column
                    if str(value).lower() == 'passed':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif str(value).lower() == 'failed':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif str(value).lower() == 'skipped':
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws_results)
        
        # Add execution summary sheet
        self._add_execution_summary_sheet(wb, execution_data, session_id)
        
        # Save workbook
        wb.save(filepath)
        return filepath
    
    def _auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths based on content"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set minimum width and maximum width limits
            adjusted_width = min(max(max_length + 2, 15), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _add_summary_sheet(self, workbook, test_cases_data, session_id):
        """Add a summary sheet with test case statistics"""
        ws_summary = workbook.create_sheet(title="Summary")
        
        # Summary headers
        ws_summary.cell(row=1, column=1, value="Test Cases Summary").font = Font(size=16, bold=True)
        ws_summary.cell(row=2, column=1, value=f"Session ID: {session_id}")
        ws_summary.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws_summary.cell(row=4, column=1, value=f"Total Test Cases: {len(test_cases_data)}")
        
        # Category breakdown
        categories = {}
        priorities = {}
        
        for test_case in test_cases_data:
            category = test_case.get('category', 'Unknown')
            priority = test_case.get('priority', 'Unknown')
            
            categories[category] = categories.get(category, 0) + 1
            priorities[priority] = priorities.get(priority, 0) + 1
        
        # Categories section
        row = 6
        ws_summary.cell(row=row, column=1, value="Categories:").font = Font(bold=True)
        for category, count in categories.items():
            row += 1
            ws_summary.cell(row=row, column=1, value=f"  {category}: {count}")
        
        # Priorities section
        row += 2
        ws_summary.cell(row=row, column=1, value="Priorities:").font = Font(bold=True)
        for priority, count in priorities.items():
            row += 1
            ws_summary.cell(row=row, column=1, value=f"  {priority}: {count}")
    
    def _add_execution_summary_sheet(self, workbook, execution_data, session_id):
        """Add a summary sheet with execution statistics"""
        ws_summary = workbook.create_sheet(title="Execution Summary")
        
        # Summary headers
        ws_summary.cell(row=1, column=1, value="Execution Summary").font = Font(size=16, bold=True)
        ws_summary.cell(row=2, column=1, value=f"Session ID: {session_id}")
        ws_summary.cell(row=3, column=1, value=f"Executed: {execution_data.get('timestamp', 'Unknown')}")
        
        # Execution statistics
        results = execution_data.get('results', [])
        total_tests = len(results)
        
        status_counts = {}
        for result in results:
            status = result.get('status', 'Unknown')
            status_counts[status] = status_counts.get(status, 0) + 1
        
        ws_summary.cell(row=5, column=1, value=f"Total Tests Executed: {total_tests}")
        
        # Status breakdown
        row = 7
        ws_summary.cell(row=row, column=1, value="Execution Results:").font = Font(bold=True)
        for status, count in status_counts.items():
            row += 1
            ws_summary.cell(row=row, column=1, value=f"  {status}: {count}")
        
        # Success rate
        passed = status_counts.get('Passed', 0)
        if total_tests > 0:
            success_rate = (passed / total_tests) * 100
            row += 2
            ws_summary.cell(row=row, column=1, value=f"Success Rate: {success_rate:.1f}%").font = Font(bold=True)